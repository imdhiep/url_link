import os
import re
import csv
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from paddleocr import PaddleOCR
from PIL import Image, ImageTk
import numpy as np
import openpyxl
import requests
from io import BytesIO

# =========================
#  ICON LOADING UTILITIES
# =========================
def load_icon_from_url(url, size=(24,24)):
    """Download icon from URL and return as Tkinter PhotoImage."""
    try:
        response = requests.get(url)
        img = Image.open(BytesIO(response.content)).convert("RGBA")
        img = img.resize(size, Image.LANCZOS)
        return ImageTk.PhotoImage(img)
    except Exception:
        return None

FOLDER_ICON_URL = "https://img.icons8.com/color/48/folder-invoices--v2.png"
START_ICON_URL = "https://img.icons8.com/color/48/circled-play.png"
RETRY_ICON_URL = "https://img.icons8.com/color/48/restart--v1.png"

# =========================
#  IMAGE PREPROCESSING
# =========================
def preprocess_image(image_path):
    img = Image.open(image_path).convert("RGB")
    w, h = img.size
    top = int(h * 0.2)
    bottom = int(h * 0.8)
    img_cropped = img.crop((0, top, w, bottom))
    return img_cropped

# =========================
#  OCR VALUE EXTRACTION
# =========================
def extract_dist1(image_path, ocr, logger=None):
    try:
        img = preprocess_image(image_path)
        arr = np.array(img)
        ocr_res = ocr.ocr(arr, cls=True)
        for line in ocr_res:
            for box, (txt, _) in line:
                txt_norm = txt.lower().replace(' ', '').replace('_', '')
                if 'dist1' in txt_norm or 'dist 1' in txt.lower().replace('_',''):
                    match = re.search(r'dist[\s_]*1[:\-]?\s*([\-\d.,]+)\s*[uμmyµ]+', txt.lower())
                    if match:
                        val = match.group(1).replace(',', '.')
                        try:
                            return float(val)
                        except:
                            continue
        flat_lines = []
        for line in ocr_res:
            for box, (txt, _) in line:
                flat_lines.append(txt)
        for idx, txt in enumerate(flat_lines):
            txt_norm = txt.lower().replace(' ', '').replace('_','')
            if 'dist1' in txt_norm or 'dist 1' in txt.lower().replace('_',''):
                if idx+1 < len(flat_lines):
                    txt2 = flat_lines[idx+1]
                    match2 = re.search(r'([\-\d.,]+)\s*[uμmyµ]+', txt2.lower())
                    if match2:
                        val2 = match2.group(1).replace(',', '.')
                        try:
                            return float(val2)
                        except:
                            continue
        for line in ocr_res:
            for box, (txt, _) in line:
                match = re.search(r'([\-\d.,]+)\s*[uμmyµ]+', txt.lower())
                if match:
                    val = match.group(1).replace(',', '.')
                    try:
                        return float(val)
                    except:
                        continue
        return 0.00
    except Exception:
        return 0.00

# =========================
#  FOLDER PROCESSING
# =========================
def process_cup_folder(folder, ocr, logger):
    results = []
    for i in range(1, 8):
        img_path = folder / f"{i}-cup.png"
        if img_path.exists():
            value = extract_dist1(img_path, ocr, logger)
            results.append({"file_name": img_path.name, "value": value})
        else:
            results.append({"file_name": f"{i}-cup.png", "value": 0.00})
    return results

def process_plunger_folder(folder, ocr, logger):
    results = []
    for group_num in range(1, 8):
        group_files = sorted(folder.glob(f"{group_num}-plunger-*.png"))
        values = []
        for img_path in group_files:
            value = extract_dist1(img_path, ocr, logger)
            results.append({"file_name": img_path.name, "value": value, "group": group_num})
            values.append((img_path.name, value))
        if values:
            max_value = max(v for _, v in values)
            for fname, v in values:
                for item in results:
                    if item["file_name"] == fname and v == max_value and max_value > 0:
                        item["highest"] = "x"
    return results

# =========================
#  EXCEL AND CSV WRITING
# =========================
def write_to_excel(cup_data, plunger_data, excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['ﾃﾞｰﾀ']
        for idx in range(7):
            row = 23 + idx * 4
            ws[f'W{row}'] = cup_data[idx]["value"]
            ws[f'X{row}'] = 0.00
        for idx in range(7):
            row = 23 + idx * 4
            group = idx+1
            group_values = [item for item in plunger_data if item.get("group", 0) == group]
            if group_values:
                max_value = max(item["value"] for item in group_values)
                ws[f'X{row}'] = max_value
        wb.save(excel_path)
        return True
    except Exception as e:
        messagebox.showerror("Excel Error", f"Cannot write file:\n{str(e)}")
        return False

def write_ocr_csv(cup_data, plunger_data, folder):
    csv_path = Path(folder) / "ocr_result.csv"
    rows = []
    for item in cup_data:
        note = "no value found" if item["value"] == 0.00 else ""
        rows.append([item["file_name"], f"{item['value']:.4f}um", "", note])
    rows.append(["", "", "", ""])
    plunger_by_group = {}
    for item in plunger_data:
        group = item.get("group", 0)
        if group not in plunger_by_group:
            plunger_by_group[group] = []
        plunger_by_group[group].append(item)
    for group in sorted(plunger_by_group.keys()):
        group_items = plunger_by_group[group]
        max_value = max(i["value"] for i in group_items) if group_items else 0
        for i in group_items:
            mark = "x" if i["value"] == max_value and max_value > 0 else ""
            note = "no value found" if i["value"] == 0.00 else ""
            rows.append([i["file_name"], f"{i['value']:.4f}um", mark, note])
    with open(csv_path, "w", newline='', encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["file name", "value", "highest_value", "note"])
        writer.writerows(rows)
    return csv_path

# =========================
#  MAIN GUI APPLICATION
# =========================
class EasyOCRApp(tk.Tk):
    """
    Main Tkinter application for Easy OCR to Excel.
    """
    def __init__(self):
        super().__init__()
        self.title("Easy OCR to Excel")
        self.geometry("540x480")
        self.resizable(False, False)
        self.lang = tk.StringVar(value="en")
        self.translations = {
            "en": {
                "title": "Easy OCR to Excel",
                "guide": (
                    "GUIDE:\n"
                    "1. Select the folder containing 'cup', 'plunger', and 'c2025.xlsx'\n"
                    "2. Images must be named: 1-cup.png, 1-plunger-1.png, ...\n"
                    "3. Click 'Start' to extract values and write to Excel\n"
                    "4. When done, paths to Excel and OCR result files will be shown"
                ),
                "select_folder": "Select Folder",
                "start": "Start",
                "processing_cup": "Processing cup images...",
                "processing_plunger": "Processing plunger images...",
                "writing_excel": "Writing to Excel...",
                "success": "Done!",
                "excel_saved": "✅ Report file saved at:",
                "ocr_saved": "✅ OCR result file saved at:",
                "error": "System error",
                "retry": "Retry"
            },
            "ja": {
                "title": "かんたんOCR→Excel",
                "guide": (
                    "使い方:\n"
                    "1. 「cup」「plunger」「c2025.xlsx」を含むフォルダを選択\n"
                    "2. 画像名は 1-cup.png, 1-plunger-1.png などにしてください\n"
                    "3. 「開始」をクリックすると値が抽出されExcelに保存されます\n"
                    "4. 完了後、ExcelとOCR結果ファイルのパスが表示されます"
                ),
                "select_folder": "フォルダ選択",
                "start": "開始",
                "processing_cup": "cup画像を処理中...",
                "processing_plunger": "plunger画像を処理中...",
                "writing_excel": "Excelに保存中...",
                "success": "完了！",
                "excel_saved": "✅ レポートファイル保存先：",
                "ocr_saved": "✅ OCR結果ファイル保存先：",
                "error": "システムエラー",
                "retry": "リトライ"
            }
        }
        # Load icons for buttons
        self.folder_icon = load_icon_from_url(FOLDER_ICON_URL)
        self.start_icon = load_icon_from_url(START_ICON_URL)
        self.retry_icon = load_icon_from_url(RETRY_ICON_URL)
        self.ocr = PaddleOCR(use_angle_cls=True, lang='en', use_gpu=False)
        self.data_folder = None
        self._setup_style()
        self.create_widgets()

    def _setup_style(self):
        style = ttk.Style(self)
        style.theme_use('clam')
        style.configure("Danger.TButton", foreground="#fff", background="#b22222", font=("Arial", 16, "bold"), borderwidth=0, focusthickness=3, focuscolor='none', padding=12)
        style.map("Danger.TButton",
                  background=[('active', '#d7263d')],
                  foreground=[('active', '#fff')])
        style.configure("Main.TFrame", background="#fff0f0")
        style.configure("TLabel", background="#fff0f0", font=("Arial", 12))
        style.configure("TEntry", fieldbackground="#fff", background="#fff0f0")
        style.configure("Custom.Horizontal.TProgressbar", thickness=22, troughcolor="#eee", background="#b22222", bordercolor="#fff0f0")

    def create_widgets(self):
        main_frame = ttk.Frame(self, padding=20, style="Main.TFrame")
        main_frame.pack(expand=True, fill='both')
        lang_frame = ttk.Frame(main_frame, style="Main.TFrame")
        lang_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(lang_frame, text="Language / 言語:", foreground="#b22222", font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        ttk.Radiobutton(lang_frame, text="English", variable=self.lang, value="en", command=self._update_language).pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(lang_frame, text="日本語", variable=self.lang, value="ja", command=self._update_language).pack(side=tk.LEFT, padx=5)
        self.title_label = ttk.Label(main_frame, text=self.translations[self.lang.get()]["title"], font=("Arial", 18, "bold"), foreground="#b22222")
        self.title_label.pack(pady=10)
        self.guide_label = tk.Label(main_frame, text=self.translations[self.lang.get()]["guide"], wraplength=480, justify="left", bg="#fff0f0", fg="#b22222", font=("Arial", 11))
        self.guide_label.pack(fill=tk.X, pady=10)
        folder_frame = ttk.Frame(main_frame, style="Main.TFrame")
        folder_frame.pack(pady=10)
        self.folder_var = tk.StringVar()
        ttk.Entry(folder_frame, textvariable=self.folder_var, width=40, state="readonly").pack(side=tk.LEFT, padx=5)
        self.btn_select = ttk.Button(folder_frame, text=self.translations[self.lang.get()]["select_folder"], image=self.folder_icon, compound="left", command=self.select_folder, style="Danger.TButton")
        self.btn_select.pack(side=tk.LEFT, padx=5)
        btns_frame = ttk.Frame(main_frame, style="Main.TFrame")
        btns_frame.pack(pady=10)
        self.btn_process = ttk.Button(btns_frame, text=self.translations[self.lang.get()]["start"], image=self.start_icon, compound="left", command=self.start_thread, state=tk.DISABLED, style="Danger.TButton")
        self.btn_process.pack(side=tk.LEFT, padx=10, ipadx=10, ipady=5)
        self.btn_retry = ttk.Button(btns_frame, text=self.translations[self.lang.get()]["retry"], image=self.retry_icon, compound="left", command=self.restart_app, style="Danger.TButton")
        self.btn_retry.pack(side=tk.LEFT, padx=10, ipadx=10, ipady=5)
        self.progress = ttk.Progressbar(main_frame, orient="horizontal", length=400, mode="determinate", style="Custom.Horizontal.TProgressbar")
        self.progress.pack(pady=10)
        self.status_label = ttk.Label(main_frame, text="", font=("Arial", 11, "bold"), foreground="#b22222")
        self.status_label.pack(pady=5)

    def _update_language(self):
        lang = self.lang.get()
        self.title(self.translations[lang]["title"])
        self.title_label.config(text=self.translations[lang]["title"])
        self.guide_label.config(text=self.translations[lang]["guide"])
        self.btn_select.config(text=self.translations[lang]["select_folder"])
        self.btn_process.config(text=self.translations[lang]["start"])
        self.btn_retry.config(text=self.translations[lang]["retry"])

    def select_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.data_folder = Path(folder)
            self.folder_var.set(str(self.data_folder))
            self.btn_process.config(state=tk.NORMAL)
            self.status_label.config(text="")

    def start_thread(self):
        self.btn_process.config(state=tk.DISABLED)
        threading.Thread(target=self.process_data, daemon=True).start()

    def restart_app(self):
        """
        Reset GUI and variables to initial state (without restarting Python process).
        """
        self.data_folder = None
        self.folder_var.set("")
        self.status_label.config(text="")
        self.progress["value"] = 0
        self.btn_process.config(state=tk.DISABLED)

    def process_data(self):
        lang = self.lang.get()
        try:
            cup_folder = self.data_folder / "cup"
            if not cup_folder.exists():
                raise FileNotFoundError(self.translations[lang]["processing_cup"])
            self.status_label.config(text=self.translations[lang]["processing_cup"])
            self.progress["value"] = 10
            self.update()
            cup_data = process_cup_folder(cup_folder, self.ocr, None)
            self.progress["value"] = 33

            plunger_folder = self.data_folder / "plunger"
            if not plunger_folder.exists():
                raise FileNotFoundError(self.translations[lang]["processing_plunger"])
            self.status_label.config(text=self.translations[lang]["processing_plunger"])
            self.update()
            plunger_data = process_plunger_folder(plunger_folder, self.ocr, None)
            self.progress["value"] = 66

            excel_path = self.data_folder / "c2025.xlsx"
            if not excel_path.exists():
                raise FileNotFoundError(self.translations[lang]["writing_excel"])
            self.status_label.config(text=self.translations[lang]["writing_excel"])
            self.update()
            ok = write_to_excel(cup_data, plunger_data, excel_path)
            csv_path = write_ocr_csv(cup_data, plunger_data, self.data_folder)
            if ok:
                self.progress["value"] = 100
                self.show_success_popup(excel_path, csv_path)
                self.status_label.config(text=self.translations[lang]["success"])
            else:
                self.status_label.config(text=self.translations[lang]["error"])
        except Exception as e:
            self.status_label.config(text=f"{self.translations[self.lang.get()]['error']}: {str(e)}", foreground="#b22222")
            tk.messagebox.showerror(self.translations[self.lang.get()]["error"], str(e))
        finally:
            self.btn_process.config(state=tk.NORMAL)

    def show_success_popup(self, excel_path, csv_path):
        lang = self.lang.get()
        popup = tk.Toplevel(self)
        popup.title(self.translations[lang]["success"])
        popup.geometry("1000x700")
        popup.resizable(False, False)
        popup.configure(bg="#fff0f0")
        ttk.Label(popup, text=self.translations[lang]["success"], font=("Arial", 26, "bold"), foreground="#b22222", background="#fff0f0", anchor="center").pack(pady=18)
        ttk.Label(popup, text=f"{self.translations[lang]['excel_saved']}", foreground="#b22222", background="#fff0f0", font=("Arial", 17, "bold"), anchor="center").pack(pady=(10,0))
        ttk.Label(popup, text=f"{excel_path.resolve()}", foreground="#333", background="#fff0f0", font=("Arial", 15), anchor="center", wraplength=640).pack(pady=(0,10))
        ttk.Label(popup, text=f"{self.translations[lang]['ocr_saved']}", foreground="#b22222", background="#fff0f0", font=("Arial", 17, "bold"), anchor="center").pack(pady=(10,0))
        ttk.Label(popup, text=f"{csv_path.resolve()}", foreground="#333", background="#fff0f0", font=("Arial", 15), anchor="center", wraplength=640).pack(pady=(0,10))
        ttk.Button(popup, text="OK", command=popup.destroy, style="Danger.TButton").pack(pady=22, ipadx=26, ipady=10)
        popup.transient(self)
        popup.grab_set()
        popup.focus_force()
        self.wait_window(popup)

# =========================
#  HOW TO USE THIS CODE
# =========================
# 1. Install requirements: pip install paddleocr pillow openpyxl requests
# 2. Run this script: python <filename>.py
# 3. Select your data folder (must contain 'cup', 'plunger', 'c2025.xlsx')
# 4. Click Start to process. Results will be saved to Excel and ocr_result.csv
# 5. Use Retry to reset and choose another folder.

if __name__ == "__main__":
    app = EasyOCRApp()
    app.mainloop()
