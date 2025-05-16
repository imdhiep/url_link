import os
import re
import csv
from pathlib import Path
from paddleocr import PaddleOCR
from PIL import Image
import numpy as np
import openpyxl
import requests
from io import BytesIO
import streamlit as st

# =========================
#  IMAGE PREPROCESSING
# =========================
def preprocess_image(image_path):
    img = Image.open(image_path).convert("RGB")
    w, h = img.size
    top = int(h * 0.2)
    bottom = int(h * 0.8)
    img_cropped = img.crop((0, top, w, bottom))
    return img_cropped

# =========================
#  OCR VALUE EXTRACTION
# =========================
def extract_dist1(image_path, ocr, logger=None):
    try:
        img = preprocess_image(image_path)
        arr = np.array(img)
        ocr_res = ocr.ocr(arr, cls=True)
        for line in ocr_res:
            for box, (txt, _) in line:
                txt_norm = txt.lower().replace(' ', '').replace('_', '')
                if 'dist1' in txt_norm or 'dist 1' in txt.lower().replace('_',''):
                    match = re.search(r'dist[\s_]*1[:\-]?\s*([\-\d.,]+)\s*[uμmyµ]+', txt.lower())
                    if match:
                        val = match.group(1).replace(',', '.')
                        try:
                            return float(val)
                        except:
                            continue
        flat_lines = []
        for line in ocr_res:
            for box, (txt, _) in line:
                flat_lines.append(txt)
        for idx, txt in enumerate(flat_lines):
            txt_norm = txt.lower().replace(' ', '').replace('_','')
            if 'dist1' in txt_norm or 'dist 1' in txt.lower().replace('_',''):
                if idx+1 < len(flat_lines):
                    txt2 = flat_lines[idx+1]
                    match2 = re.search(r'([\-\d.,]+)\s*[uμmyµ]+', txt2.lower())
                    if match2:
                        val2 = match2.group(1).replace(',', '.')
                        try:
                            return float(val2)
                        except:
                            continue
        for line in ocr_res:
            for box, (txt, _) in line:
                match = re.search(r'([\-\d.,]+)\s*[uμmyµ]+', txt.lower())
                if match:
                    val = match.group(1).replace(',', '.')
                    try:
                        return float(val)
                    except:
                        continue
        return 0.00
    except Exception:
        return 0.00

# =========================
#  FOLDER PROCESSING
# =========================
def process_cup_folder(folder, ocr, logger):
    results = []
    for i in range(1, 8):
        img_path = folder / f"{i}-cup.png"
        if img_path.exists():
            value = extract_dist1(img_path, ocr, logger)
            results.append({"file_name": img_path.name, "value": value})
        else:
            results.append({"file_name": f"{i}-cup.png", "value": 0.00})
    return results

def process_plunger_folder(folder, ocr, logger):
    results = []
    for group_num in range(1, 8):
        group_files = sorted(folder.glob(f"{group_num}-plunger-*.png"))
        values = []
        for img_path in group_files:
            value = extract_dist1(img_path, ocr, logger)
            results.append({"file_name": img_path.name, "value": value, "group": group_num})
            values.append((img_path.name, value))
        if values:
            max_value = max(v for _, v in values)
            for fname, v in values:
                for item in results:
                    if item["file_name"] == fname and v == max_value and max_value > 0:
                        item["highest"] = "x"
    return results

# =========================
#  EXCEL AND CSV WRITING
# =========================
def write_to_excel(cup_data, plunger_data, excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['ﾃﾞｰﾀ']
        for idx in range(7):
            row = 23 + idx * 4
            ws[f'W{row}'] = cup_data[idx]["value"]
            ws[f'X{row}'] = 0.00
        for idx in range(7):
            row = 23 + idx * 4
            group = idx+1
            group_values = [item for item in plunger_data if item.get("group", 0) == group]
            if group_values:
                max_value = max(item["value"] for item in group_values)
                ws[f'X{row}'] = max_value
        wb.save(excel_path)
        return True
    except Exception as e:
        st.error(f"Excel Error: Cannot write file: {str(e)}")
        return False

def write_ocr_csv(cup_data, plunger_data, folder):
    csv_path = Path(folder) / "ocr_result.csv"
    rows = []
    for item in cup_data:
        note = "no value found" if item["value"] == 0.00 else ""
        rows.append([item["file_name"], f"{item['value']:.4f}um", "", note])
    rows.append(["", "", "", ""])
    plunger_by_group = {}
    for item in plunger_data:
        group = item.get("group", 0)
        if group not in plunger_by_group:
            plunger_by_group[group] = []
        plunger_by_group[group].append(item)
    for group in sorted(plunger_by_group.keys()):
        group_items = plunger_by_group[group]
        max_value = max(i["value"] for i in group_items) if group_items else 0
        for i in group_items:
            mark = "x" if i["value"] == max_value and max_value > 0 else ""
            note = "no value found" if i["value"] == 0.00 else ""
            rows.append([i["file_name"], f"{i['value']:.4f}um", mark, note])
    with open(csv_path, "w", newline='', encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["file name", "value", "highest_value", "note"])
        writer.writerows(rows)
    return csv_path

# =========================
#  MAIN STREAMLIT APPLICATION
# =========================
def main():
    st.set_page_config(page_title="Easy OCR to Excel", layout="centered")
    lang = st.radio("Language / 言語:", ["en", "ja"], horizontal=True)
    translations = {
        "en": {
            "title": "Easy OCR to Excel",
            "guide": (
                "GUIDE:\n"
                "1. Enter the folder path containing 'cup', 'plunger', and 'c2025.xlsx'\n"
                "2. Images must be named: 1-cup.png, 1-plunger-1.png, ...\n"
                "3. Click 'Start' to extract values and write to Excel\n"
                "4. When done, paths to Excel and OCR result files will be shown"
            ),
            "select_folder": "Folder Path",
            "start": "Start",
            "processing_cup": "Processing cup images...",
            "processing_plunger": "Processing plunger images...",
            "writing_excel": "Writing to Excel...",
            "success": "Done!",
            "excel_saved": "✅ Report file saved at:",
            "ocr_saved": "✅ OCR result file saved at:",
            "error": "System error",
            "retry": "Retry"
        },
        "ja": {
            "title": "かんたんOCR→Excel",
            "guide": (
                "使い方:\n"
                "1. 「cup」「plunger」「c2025.xlsx」を含むフォルダのパスを入力\n"
                "2. 画像名は 1-cup.png, 1-plunger-1.png などにしてください\n"
                "3. 「開始」をクリックすると値が抽出されExcelに保存されます\n"
                "4. 完了後、ExcelとOCR結果ファイルのパスが表示されます"
            ),
            "select_folder": "フォルダパス",
            "start": "開始",
            "processing_cup": "cup画像を処理中...",
            "processing_plunger": "plunger画像を処理中...",
            "writing_excel": "Excelに保存中...",
            "success": "完了！",
            "excel_saved": "✅ レポートファイル保存先：",
            "ocr_saved": "✅ OCR結果ファイル保存先：",
            "error": "システムエラー",
            "retry": "リトライ"
        }
    }
    st.title(translations[lang]["title"])
    st.info(translations[lang]["guide"])
    folder = st.text_input(translations[lang]["select_folder"], value="")
    start_btn = st.button(translations[lang]["start"])
    progress = st.empty()
    status = st.empty()
    if start_btn and folder:
        try:
            cup_folder = Path(folder) / "cup"
            if not cup_folder.exists():
                st.error(translations[lang]["processing_cup"])
                return
            progress.progress(10, text=translations[lang]["processing_cup"])
            ocr = PaddleOCR(use_angle_cls=True, lang='en', use_gpu=False)
            cup_data = process_cup_folder(cup_folder, ocr, None)
            progress.progress(33, text=translations[lang]["processing_plunger"])
            plunger_folder = Path(folder) / "plunger"
            if not plunger_folder.exists():
                st.error(translations[lang]["processing_plunger"])
                return
            plunger_data = process_plunger_folder(plunger_folder, ocr, None)
            progress.progress(66, text=translations[lang]["writing_excel"])
            excel_path = Path(folder) / "c2025.xlsx"
            if not excel_path.exists():
                st.error(translations[lang]["writing_excel"])
                return
            ok = write_to_excel(cup_data, plunger_data, excel_path)
            csv_path = write_ocr_csv(cup_data, plunger_data, folder)
            if ok:
                progress.progress(100, text=translations[lang]["success"])
                st.success(translations[lang]["success"])
                st.write(f"{translations[lang]['excel_saved']} {excel_path.resolve()}")
                st.write(f"{translations[lang]['ocr_saved']} {csv_path.resolve()}")
            else:
                st.error(translations[lang]["error"])
        except Exception as e:
            st.error(f"{translations[lang]['error']}: {str(e)}")

if __name__ == "__main__":
    main()
